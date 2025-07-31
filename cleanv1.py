import streamlit as st
import openpyxl
import io
import pandas as pd
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel Cleaner & Comparator", layout="wide")
st.title("📊 Excel Cleaner & Comparator")

@st.cache_data
def get_headers(file):
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active
    max_col = sheet.max_column
    headers = [
        sheet.cell(row=1, column=col).value or f"(empty {get_column_letter(col)})"
        for col in range(1, max_col + 1)
    ]
    return headers, max_col, sheet.max_row

# ---------------------------
# Step 1: Clean and Download Original File
# ---------------------------
st.subheader("Step 1: Upload, Clean, and Save the Original File")
original_file = st.file_uploader("Upload Original Excel File", type=["xlsx", "xlsm"], key="original")

if original_file:
    headers_orig, max_col_orig, max_row_orig = get_headers(original_file)
    st.write(f"✅ Original file has **{max_col_orig} columns**")

    selected_cols_orig = st.multiselect(
        "Select columns to delete from the Original File:",
        options=list(range(1, max_col_orig + 1)),
        format_func=lambda x: f"{x}: {headers_orig[x-1]}",
        key="delete_orig"
    )
    st.info(f"Selected {len(selected_cols_orig)} / {max_col_orig} columns for deletion")

    if st.button("Clean & Download Original File"):
        wb_orig = openpyxl.load_workbook(original_file, read_only=True, data_only=True)
        sheet_orig = wb_orig.active
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = f"{sheet_orig.title}_cleaned"

        # Build new headers excluding deleted columns
        cleaned_headers = [h for i, h in enumerate(headers_orig) if i+1 not in selected_cols_orig]

        for row in sheet_orig.iter_rows(values_only=True, max_col=max_col_orig):
            new_row = [cell for idx, cell in enumerate(row, start=1) if idx not in selected_cols_orig]
            new_sheet.append(new_row)

        output = io.BytesIO()
        new_wb.save(output)
        output.seek(0)

        st.success("✅ Original File cleaned and ready for download.")
        st.download_button(
            label="Download Cleaned Original File",
            data=output,
            file_name="cleaned_original.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Save headers in session state for Step 2
        st.session_state["cleaned_headers"] = cleaned_headers

# ---------------------------
# Step 2: Upload and Match New File
# ---------------------------
if "cleaned_headers" in st.session_state:
    st.subheader("Step 2: Upload New File to Match Cleaned Original")
    new_file = st.file_uploader("Upload New Excel File", type=["xlsx", "xlsm"], key="new")

    if new_file:
        headers_new, max_col_new, max_row_new = get_headers(new_file)
        st.write(f"✅ New file has **{max_col_new} columns**")

        cleaned_headers = st.session_state["cleaned_headers"]

        # Compare headers
        common_headers = [h for h in cleaned_headers if h in headers_new]
        missing_in_new = [h for h in cleaned_headers if h not in headers_new]
        extra_in_new = [h for h in headers_new if h not in cleaned_headers]

        st.subheader("🔍 Comparison Results")
        st.write(f"✅ Columns in common: {len(common_headers)}")
        if missing_in_new:
            st.warning(f"⚠️ Missing in new file: {missing_in_new}")
        if extra_in_new:
            st.warning(f"⚠️ Extra in new file: {extra_in_new}")

        if st.button("Preview Matched New File"):
            wb_preview_new = openpyxl.load_workbook(new_file, read_only=True, data_only=True)
            sheet_preview_new = wb_preview_new.active
            df_new = pd.DataFrame(sheet_preview_new.iter_rows(values_only=True, max_col=max_col_new))

            keep_indices = [headers_new.index(h) for h in common_headers]
            df_new_clean = df_new.iloc[:, keep_indices]
            df_new_clean.columns = common_headers

            st.subheader("Preview of Matched New File (first 10 rows)")
            st.dataframe(df_new_clean.head(10))

        if st.button("Clean and Download Matched New File"):
            wb_new = openpyxl.load_workbook(new_file, read_only=True, data_only=True)
            sheet_new = wb_new.active
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            new_sheet.title = f"{sheet_new.title}_matched"

            keep_indices = [headers_new.index(h) for h in common_headers]

            for row in sheet_new.iter_rows(values_only=True, max_col=max_col_new):
                new_row = [row[i] for i in keep_indices]
                new_sheet.append(new_row)

            output = io.BytesIO()
            new_wb.save(output)
            output.seek(0)

            st.success("✅ New file matched to Cleaned Original successfully!")
            st.download_button(
                label="Download Matched Excel File",
                data=output,
                file_name="matched_new_file.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
